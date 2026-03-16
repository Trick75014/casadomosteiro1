"""
convert.py — Converte qualquer ficheiro .xlsx → reservas.json
Executado automaticamente pelo GitHub Actions.
"""
import json, glob, sys, os, re
from datetime import datetime, date

# ── Encontrar o ficheiro Excel ────────────────────────────────────────
files = glob.glob('*.xlsx') + glob.glob('*.xls')
if not files:
    print("ERRO: Nenhum ficheiro Excel encontrado!")
    print("Ficheiros presentes:", os.listdir('.'))
    sys.exit(1)

xlsx = sorted(files, key=os.path.getmtime, reverse=True)[0]
print(f"Ficheiro: {xlsx}")

try:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, data_only=True)
except Exception as e:
    print(f"ERRO ao abrir Excel: {e}")
    sys.exit(1)

print(f"Folhas: {wb.sheetnames}")

# ── Folha de reservas ─────────────────────────────────────────────────
ws = None
for name in wb.sheetnames:
    if any(x in name.lower() for x in ['serva', 'reserv']):
        ws = wb[name]; break
if ws is None:
    ws = wb.active
print(f"Folha: {ws.title}")

# ── Cabeçalho ─────────────────────────────────────────────────────────
header_row = None
for r in range(1, 8):
    vals = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, 20))
    if any(x in vals for x in ['Voyageur','voyageur','Nome','nome']):
        header_row = r; break

if not header_row:
    print("ERRO: Cabeçalho nao encontrado!")
    for r in range(1,6):
        print(f"Linha {r}:", [ws.cell(r,c).value for c in range(1,10)])
    sys.exit(1)

print(f"Cabecalho linha {header_row}")

# ── Colunas ───────────────────────────────────────────────────────────
def norm(s):
    import unicodedata
    s = str(s or '').strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

hdr = {norm(ws.cell(header_row, c).value): c for c in range(1, 20) if ws.cell(header_row, c).value}

def fc(*keys):
    for k in keys:
        kn = norm(k)
        for h, c in hdr.items():
            if kn in h or h in kn:
                return c
    return None

iPlat  = fc('Plateforme','Plataforma')
iName  = fc('Voyageur','Nome','Guest')
iIn    = fc('Arrivee','Arrivee','Chegada','Checkin','Check-in')
iOut   = fc('Depart','Saida','Checkout','Check-out')
iG     = fc('Personnes','Pessoas','Guests')
iHour  = fc('Hora','Heure')
iPhone = fc('Telephone','Telef','Phone')

print(f"Cols: name={iName} in={iIn} out={iOut} plat={iPlat} guests={iG} hour={iHour} phone={iPhone}")

if not iName or not iIn or not iOut:
    print("ERRO: Colunas essenciais nao encontradas!")
    sys.exit(1)

# ── Datas ─────────────────────────────────────────────────────────────
def ymd(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})', s)
    if m: return f"{m[3]}-{m[2]}-{m[1]}"
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m: return s[:10]
    return None

# ── Reservas ──────────────────────────────────────────────────────────
R = []
for r in range(header_row + 1, ws.max_row + 1):
    name = ws.cell(r, iName).value
    if not name: continue
    ns = str(name).strip()
    if not ns or ns in ('TOTAUX','TOTAL','Total'): continue
    ci = ymd(ws.cell(r, iIn).value)
    co = ymd(ws.cell(r, iOut).value)
    if not ci or not co: continue
    g = ws.cell(r, iG).value if iG else 1
    try: guests = int(float(str(g))) if g else 1
    except: guests = 1
    R.append({
        "name":     ns,
        "checkin":  ci,
        "checkout": co,
        "platform": str(ws.cell(r,iPlat).value or '').strip() if iPlat else '',
        "guests":   guests,
        "hour":     str(ws.cell(r,iHour).value or '').strip() if iHour else '',
        "phone":    str(ws.cell(r,iPhone).value or '').strip() if iPhone else '',
    })
    print(f"  + {R[-1]['platform']} | {ns} | {ci} -> {co}")

print(f"\nTotal: {len(R)} reservas")
if not R:
    print("Nenhuma reserva — sem alteracoes")
    sys.exit(0)

with open('reservas.json', 'w', encoding='utf-8') as f:
    json.dump(R, f, ensure_ascii=False, indent=2)
print("reservas.json guardado!")
